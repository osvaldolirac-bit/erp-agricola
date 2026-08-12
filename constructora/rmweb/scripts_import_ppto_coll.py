#!/usr/bin/env python3
"""Importa ppto+APU Excel a una obra demo (nombres ficticios).

Soporta dos layouts de presupuesto:
  A) ITEM | PARTIDA | NOTAS | UNIDAD | CANTIDAD | PRECIO | TOTAL
  B) Ítem | Partida | Unidad | Cantidad | Precio | Total
"""
from __future__ import annotations

import argparse
import re
import sqlite3
import unicodedata
import urllib.request
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

from rmweb import constructora as cst
from rmweb import core
from rmweb import obra_contrato as obractx

DB_DEFAULT = "/root/constructora/data/constructora_demo.db"


def _f(v, default=0.0):
    if v is None or v == "":
        return float(default)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace("%", "").replace(" ", "")
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return float(default)


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def pick_ppto_sheet(wb):
    for name in ("Ppto", "Casa A", "Presupuesto", "Cotizacion", "Cotización"):
        if name in wb.sheetnames:
            return wb[name]
    # first sheet that looks like ppto
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            joined = " ".join(str(v).lower() for v in row if v is not None)
            if "partida" in joined and ("ítem" in joined or "item" in joined):
                return ws
    return wb[wb.sheetnames[0]]


def pick_apu_sheet(wb):
    for name in wb.sheetnames:
        if "apu" in name.lower().replace(".", ""):
            return wb[name]
    return None


def detect_layout(ws):
    """Return 'A' (con notas) o 'B' (sin notas)."""
    for row in ws.iter_rows(values_only=True):
        cells = [("" if v is None else str(v)).strip().lower() for v in list(row)[:7]]
        if not cells:
            continue
        if cells[0] in ("item", "ítem", "ìtem") or cells[0].startswith("ítem") or cells[0].startswith("item"):
            # B: item, partida, unidad...
            if len(cells) > 2 and cells[2] in ("unidad", "und", "un"):
                return "B"
            # A: item, partida, notas, unidad
            if len(cells) > 3 and cells[3] in ("unidad", "und", "un"):
                return "A"
            if "partida" in cells[1]:
                # heuristic: if col2 looks like unidad header
                if "unidad" in cells[2]:
                    return "B"
                return "A"
    return "B"


def parse_ppto(ws):
    layout = detect_layout(ws)
    header = {
        "ubicacion": None,
        "propietario": None,
        "documento_cot": None,
        "duracion_meses": 10.0,
        "valor_uf": 0.0,
        "gg_pct": 5.0,
        "utilidades_pct": 21.0,
        "iva_pct": 19.0,
    }
    for row in ws.iter_rows(values_only=True):
        a = ("" if row[0] is None else str(row[0])).strip().lower()
        b = row[1] if len(row) > 1 else None
        if a.startswith("ubic"):
            header["ubicacion"] = ("" if b is None else str(b)).strip()
        elif a.startswith("mandant") or a.startswith("propiet"):
            header["propietario"] = ("" if b is None else str(b)).strip()
        elif a.startswith("documento"):
            header["documento_cot"] = ("" if b is None else str(b)).strip()
        elif a.startswith("duraci"):
            m = re.search(r"([\d.,]+)", str(b or ""))
            if m:
                header["duracion_meses"] = _f(m.group(1))
        # UF value often near footer label
        for idx, v in enumerate(row):
            if v is None:
                continue
            s = str(v).lower()
            if "valor u.f" in s or "valor uf" in s:
                for vv in row[idx + 1 :]:
                    if isinstance(vv, (int, float)) and float(vv) > 1000:
                        header["valor_uf"] = float(vv)
                        break
                    fv = _f(vv)
                    if fv > 1000:
                        header["valor_uf"] = fv
                        break
        # footer GG/util/IVA — label may be col B or C
        for idx in (2, 3, 1):
            if len(row) <= idx or row[idx] is None:
                continue
            lab = str(row[idx]).strip().lower()
            if lab == "gastos generales":
                # pct usually next numeric after label
                for vv in row[idx + 1 :]:
                    if vv is None or vv == "":
                        continue
                    header["gg_pct"] = _f(vv) * (100.0 if _f(vv) <= 1 else 1.0)
                    break
            elif lab == "utilidades":
                for vv in row[idx + 1 :]:
                    if vv is None or vv == "":
                        continue
                    header["utilidades_pct"] = _f(vv) * (100.0 if _f(vv) <= 1 else 1.0)
                    break
            elif lab == "iva":
                for vv in row[idx + 1 :]:
                    if vv is None or vv == "":
                        continue
                    header["iva_pct"] = _f(vv) * (100.0 if _f(vv) <= 1 else 1.0)
                    break

    lines = []
    started = False
    for row in ws.iter_rows(values_only=True):
        cells = list(row) + [None] * 8
        item = cells[0]
        if item is not None and str(item).strip().lower() in ("item", "ítem", "ìtem"):
            started = True
            continue
        if not started or item is None or str(item).strip() == "":
            continue
        if layout == "A":
            partida, notas, und, cant, pu = cells[1], cells[2], cells[3], cells[4], cells[5]
        else:
            partida, notas, und, cant, pu = cells[1], "", cells[2], cells[3], cells[4]

        lab = ("" if und is None else str(und)).strip().lower()
        if lab in ("sub total", "gastos generales", "utilidades", "neto", "iva", "total"):
            continue
        # footer sometimes puts label in partida/unidad cols
        if str(partida or "").strip().lower() in (
            "sub total",
            "gastos generales",
            "utilidades",
            "neto",
            "iva",
            "total",
        ):
            continue

        codigo = str(item).strip()
        detalle = ("" if partida is None else str(partida)).strip()
        if not detalle:
            continue
        notas_s = ("" if notas is None else str(notas)).strip()
        und_s = ("" if und is None else str(und)).strip()
        cant_f = _f(cant)
        pu_f = _f(pu)

        marca = ""
        blob = f"{notas_s} {detalle}".lower()
        if "gastos generales" in blob:
            marca = "en_gg"
        elif "mandante" in blob:
            marca = "mandante"
        elif "a definir" in blob or notas_s.lower() == "n/a":
            marca = "a_definir"
        elif "proforma" in blob:
            marca = "proforma"
        elif "sub-contrato" in und_s.lower() or "subcontrato" in und_s.lower():
            marca = "subcontrato"
            und_s = "gl"

        es_cap = False
        if re.fullmatch(r"\d+\.0", codigo) and not und_s:
            es_cap = True
        elif not und_s and pu_f <= 0 and not marca:
            es_cap = True

        # normalizar unidad
        und_map = {"unidad": "uni", "mt2": "m2", "mt3": "m3", "m²": "m2", "m³": "m3"}
        und_s = und_map.get(und_s.lower(), und_s) if und_s else und_s

        lines.append(
            {
                "codigo": codigo,
                "detalle": detalle,
                "notas": notas_s,
                "unidad": "" if es_cap else (und_s or "gl"),
                "cantidad": 0.0
                if es_cap
                else (cant_f if cant_f > 0 else (1.0 if pu_f > 0 else 0.0)),
                "pu": 0.0
                if es_cap or marca in ("en_gg", "mandante", "a_definir")
                else pu_f,
                "tipo_linea": "capitulo" if es_cap else "partida",
                "marca": marca,
            }
        )
    return header, lines, layout


def parse_apu_blocks(ws):
    if ws is None:
        return []
    blocks = []
    current = None
    for row in ws.iter_rows(values_only=True):
        vals = [("" if v is None else str(v).strip()) for v in list(row)[:6]]
        a, b, c, d, e, f = (vals + [""] * 6)[:6]
        if a and not b and not c and a.lower() not in ("nº", "n°", "no"):
            # skip banner
            if "detalle de la actividad" in a.lower():
                continue
            if current and current["items"]:
                blocks.append(current)
            current = {
                "titulo": a,
                "items": [],
                "leyes_pct": 29.0,
                "perdidas_pct": 10.0,
                "pu": None,
            }
            continue
        if not current:
            continue
        if a.lower() in ("nº", "n°", "nº.", "no") or b.lower() == "especificación":
            continue
        label = (d or b or "").lower()
        if "total general" in label:
            current["pu"] = _f(f or e)
            continue
        if any(
            k in label
            for k in ("materiales", "mano de obra", "sub contrato", "subcontrato")
        ):
            continue
        if not a or not b:
            continue
        bl = b.lower()
        if "leyes sociales" in bl:
            current["leyes_pct"] = _f(c) * (100.0 if _f(c) <= 1 else 1.0)
            continue
        if "perdida" in bl or "pérdida" in bl:
            current["perdidas_pct"] = _f(c) * (100.0 if _f(c) <= 1 else 1.0)
            continue
        und = d or "un"
        cant = _f(c)
        precio = _f(e)
        if cant <= 0 and precio <= 0:
            continue
        tipo = "insumo"
        if any(
            k in bl
            for k in (
                "cuadrilla",
                "maestro",
                "ayudante",
                "jornal",
                "soldador",
                "mano de obra",
                "personas",
                "direccion",
                "dirección",
            )
        ) or und.lower() in ("jh", "hh"):
            tipo = "mano_obra"
        elif any(k in bl for k in ("arriendo", "flete", "excavadora", "rodillo", "camión", "camion")):
            tipo = "otro"
        current["items"].append(
            {
                "descripcion": b,
                "cantidad": cant if cant > 0 else 1.0,
                "unidad": und,
                "precio_unitario": precio,
                "tipo": tipo,
            }
        )
    if current and current["items"]:
        blocks.append(current)
    return blocks


def best_apu(partida_nombre: str, blocks, used: set[int]):
    pn = norm(partida_nombre)
    best_i, best_s = None, 0.0
    for i, b in enumerate(blocks):
        if i in used:
            continue
        bn = norm(b["titulo"])
        if not bn:
            continue
        score = SequenceMatcher(None, pn, bn).ratio()
        if bn in pn or pn in bn:
            score = max(score, 0.85)
        ta, tb = set(pn.split()), set(bn.split())
        if ta and tb:
            score = max(score, len(ta & tb) / max(len(ta), len(tb)))
        if score > best_s:
            best_s, best_i = score, i
    if best_i is not None and best_s >= 0.55:
        return best_i, best_s
    return None, 0.0


def wipe_obras(c):
    rows = c.execute(
        "SELECT id FROM centros_costo WHERE COALESCE(tipo,'general')='obra'"
    ).fetchall()
    ids = [int(r["id"]) for r in rows]
    for oid in ids:
        for a in c.execute("SELECT id FROM apu WHERE centro_costo_id=?", (oid,)):
            c.execute("DELETE FROM apu_items WHERE apu_id=?", (int(a["id"]),))
        c.execute("DELETE FROM apu WHERE centro_costo_id=?", (oid,))
        c.execute("DELETE FROM obra_partidas WHERE centro_costo_id=?", (oid,))
        for sql in (
            "DELETE FROM obra_subcontratos WHERE obra_id=?",
            "DELETE FROM obra_subcontratos WHERE centro_costo_id=?",
            "DELETE FROM obra_eepp_items WHERE eepp_id IN (SELECT id FROM obra_eepp WHERE centro_costo_id=?)",
            "DELETE FROM obra_eepp WHERE centro_costo_id=?",
        ):
            try:
                c.execute(sql, (oid,))
            except sqlite3.Error:
                pass
        try:
            for cot in c.execute(
                "SELECT id FROM cotizaciones WHERE centro_costo_id=?", (oid,)
            ):
                c.execute(
                    "DELETE FROM cotizacion_items WHERE cotizacion_id=?",
                    (int(cot["id"]),),
                )
            c.execute("DELETE FROM cotizaciones WHERE centro_costo_id=?", (oid,))
        except sqlite3.Error:
            pass
        try:
            c.execute(
                "DELETE FROM bodegas WHERE obra_id=? OR centro_costo_id=?", (oid, oid)
            )
        except sqlite3.Error:
            pass
        c.execute("DELETE FROM centros_costo WHERE id=?", (oid,))
    return ids


def import_obra(
    *,
    xlsx: Path,
    nombre: str,
    ubicacion: str,
    propietario: str,
    wipe: bool,
    db_path: str,
    url: str | None = None,
):
    if (not xlsx.exists() or xlsx.stat().st_size < 1000) and url:
        urllib.request.urlretrieve(url, xlsx)

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws_ppto = pick_ppto_sheet(wb)
    ws_apu = pick_apu_sheet(wb)
    header, lines, layout = parse_ppto(ws_ppto)
    blocks = parse_apu_blocks(ws_apu)
    print(
        f"Sheet ppto={ws_ppto.title} layout={layout} líneas={len(lines)} | "
        f"APU={ws_apu.title if ws_apu else '-'} blocks={len(blocks)}"
    )

    c = sqlite3.connect(db_path)
    c.row_factory = sqlite3.Row
    cst.ensure_constructora_schema(c)
    obractx.ensure_obra_contrato_schema(c)

    if wipe:
        deleted = wipe_obras(c)
        print("Obras eliminadas:", deleted)

    # si ya existe el nombre, agregar sufijo
    nom = nombre
    exists = c.execute(
        "SELECT id FROM centros_costo WHERE lower(nombre)=lower(?)", (nom,)
    ).fetchone()
    if exists:
        nom = f"{nombre} ({core.hoy_chile().isoformat()})"

    ok, msg, oid = cst.crear_obra(c, nombre=nom, presupuesto=0, notas_estado="activa")
    if not ok:
        raise SystemExit(msg)
    print(msg, "id=", oid)

    obractx.guardar_parametros_presupuesto(
        c,
        oid,
        ubicacion=ubicacion,
        propietario=propietario,
        documento_cot=header.get("documento_cot") or "Presupuesto preliminar",
        duracion_meses=header.get("duracion_meses") or 10,
        gg_pct=header.get("gg_pct") or 5,
        utilidades_pct=header.get("utilidades_pct") or 21,
        descuento_clp=0,
        iva_pct=header.get("iva_pct") or 19,
        valor_uf=header.get("valor_uf") or 0,
    )
    c.execute(
        "UPDATE centros_costo SET cotizacion_obra_estado='borrador' WHERE id=?",
        (oid,),
    )

    hoy = core.hoy_chile().isoformat()
    pid_by_orden = {}
    n_cap = n_part = 0
    for orden, ln in enumerate(lines, start=1):
        cur = c.execute(
            """
            INSERT INTO obra_partidas
            (centro_costo_id, codigo, detalle, unidad, cantidad, pu_neto, total,
             avance_pct, orden, notas, activo, creado_en, tipo_linea, marca, apu_id)
            VALUES (?,?,?,?,?,0,0,0,?,?,1,?,?,?,NULL)
            """,
            (
                oid,
                ln["codigo"],
                ln["detalle"],
                ln["unidad"],
                ln["cantidad"],
                orden,
                ln["notas"] or None,
                hoy,
                ln["tipo_linea"],
                ln["marca"] or None,
            ),
        )
        pid_by_orden[orden] = int(cur.lastrowid)
        if ln["tipo_linea"] == "capitulo":
            n_cap += 1
        else:
            n_part += 1

    used_apu: set[int] = set()
    n_apu = n_stub = 0
    for orden, ln in enumerate(lines, start=1):
        if ln["tipo_linea"] == "capitulo":
            continue
        if ln["marca"] in ("en_gg", "mandante", "a_definir"):
            continue
        pid = pid_by_orden[orden]
        bi, score = best_apu(ln["detalle"], blocks, used_apu)
        items = []
        leyes = 0.0
        perdidas = 0.0
        if bi is not None:
            used_apu.add(bi)
            blk = blocks[bi]
            items = blk["items"]
            leyes = float(blk.get("leyes_pct") or 0)
            perdidas = float(blk.get("perdidas_pct") or 0)
            n_apu += 1
            notas = f"Import APU «{blk['titulo']}» score={score:.2f}"
        elif ln["pu"] > 0:
            items = [
                {
                    "descripcion": ln["detalle"],
                    "cantidad": 1.0,
                    "unidad": ln["unidad"] or "gl",
                    "precio_unitario": ln["pu"],
                    "tipo": "otro",
                }
            ]
            n_stub += 1
            notas = "PU desde ppto (sin desglose APU en planilla)"
        else:
            continue

        ok, msg, apu_id = cst.guardar_apu(
            c,
            apu_id=None,
            obra_id=oid,
            codigo="",
            nombre=ln["detalle"][:120],
            unidad=ln["unidad"] or "gl",
            leyes_pct=leyes,
            perdidas_pct=perdidas,
            notas=notas,
            activo=1,
            items=items,
            partida_id=pid,
        )
        if not ok:
            print("APU fail", ln["codigo"], ln["detalle"][:40], msg)
            continue
        if ln["pu"] > 0:
            c.execute(
                """
                UPDATE obra_partidas
                SET pu_neto=?, total=ROUND(?*?,2), cantidad=?, unidad=?
                WHERE id=?
                """,
                (
                    ln["pu"],
                    ln["cantidad"],
                    ln["pu"],
                    ln["cantidad"],
                    ln["unidad"] or "gl",
                    pid,
                ),
            )
            c.execute(
                "UPDATE apu SET pu_neto=? WHERE id=? AND centro_costo_id=?",
                (ln["pu"], int(apu_id), oid),
            )
        else:
            c.execute(
                "UPDATE obra_partidas SET pu_neto=0, total=0 WHERE id=?", (pid,)
            )
            c.execute(
                "UPDATE apu SET pu_neto=0 WHERE id=? AND centro_costo_id=?",
                (int(apu_id), oid),
            )

    tot = obractx.totales_cotizacion_obra(c, oid)
    c.execute(
        "UPDATE centros_costo SET presupuesto=? WHERE id=?",
        (float(tot.get("neto") or 0), oid),
    )
    c.commit()
    print(
        f"OK obra {oid}: caps={n_cap} partidas={n_part} apu_match={n_apu} apu_stub={n_stub}"
    )
    print(
        "Totales:",
        {
            k: tot.get(k)
            for k in ("subtotal", "gg", "utilidades", "neto", "iva", "total", "monto_uf")
        },
    )
    print(
        "Obras:",
        [dict(r) for r in c.execute("select id,nombre from centros_costo where tipo='obra' order by id")],
    )
    c.close()
    return oid


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--url", default="")
    ap.add_argument("--nombre", required=True)
    ap.add_argument("--ubicacion", default="Ubicación demo")
    ap.add_argument("--propietario", default="Propietario demo")
    ap.add_argument("--wipe", action="store_true", help="Elimina todas las obras antes")
    ap.add_argument("--db", default=DB_DEFAULT)
    args = ap.parse_args()
    import_obra(
        xlsx=Path(args.xlsx),
        nombre=args.nombre,
        ubicacion=args.ubicacion,
        propietario=args.propietario,
        wipe=bool(args.wipe),
        db_path=args.db,
        url=args.url or None,
    )


if __name__ == "__main__":
    main()
