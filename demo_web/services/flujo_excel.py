"""Exportación Excel del flujo mensual con fórmulas."""
from __future__ import annotations

import io
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_DISPLAY_COLS = [
    ("MES", "MES"),
    ("INGRESOS", "INGRESOS"),
    ("RRHH", "RRHH SUELDOS"),
    ("TESO_REAL", "TESO REAL"),
    ("TESO_PROY", "TESO PROY"),
    ("EGRESOS_REAL", "EGRESOS REAL"),
    ("EGRESOS_PROY", "EGRESOS PROY"),
    ("EGRESOS_TOTAL", "EGRESOS TOTAL"),
    ("RESULTADO_MES", "RESULTADO MES"),
    ("EERR_ACUM", "EERR ACUM"),
]

# Índices 1-based en hoja (A=1 …): F=6 EG REAL, G=7 EG PROY, H=8 TOTAL, B=2 ING, I=9 RES
_COL_EG_REAL = 6
_COL_EG_PROY = 7
_COL_EG_TOTAL = 8
_COL_ING = 2
_COL_RES = 9


def _safe_name(text: str) -> str:
    out = "".join(c if c.isalnum() or c in "._-" else "_" for c in (text or "flujo"))
    return out[:80] or "flujo"


def build_flujo_mensual_xlsx(
    df_flujo: pd.DataFrame,
    *,
    temporada: str,
    fi: date,
    ff: date,
    fila_total: dict | None = None,
) -> bytes:
    """Genera planilla Excel del flujo mensual. Totales y EERR con fórmulas donde aplica."""
    from erp_flujo_financiero import fila_total_flujo_mensual

    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo mensual"

    ws["A1"] = f"Flujo financiero — {temporada}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Temporada {fi.strftime('%d-%m-%Y')} → {ff.strftime('%d-%m-%Y')}"
    ws["A2"].font = Font(size=10, color="666666")

    header_row = 4
    for col_idx, (_, label) in enumerate(_DISPLAY_COLS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1565C0")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if df_flujo is None or df_flujo.empty:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    data_start = header_row + 1
    base_keys = [k for k, _ in _DISPLAY_COLS]
    n_rows = len(df_flujo)

    for i, (_, row) in enumerate(df_flujo.iterrows()):
        r = data_start + i
        for col_idx, key in enumerate(base_keys, start=1):
            if key == "MES":
                ws.cell(row=r, column=col_idx, value=str(row.get(key) or ""))
                continue
            if key in {"EGRESOS_TOTAL", "RESULTADO_MES"}:
                continue
            val = float(row.get(key) or 0)
            ws.cell(row=r, column=col_idx, value=val).number_format = '#,##0'

        ws.cell(row=r, column=_COL_EG_TOTAL, value=f"=F{r}+G{r}").number_format = '#,##0'
        ws.cell(row=r, column=_COL_RES, value=f"=B{r}-H{r}").number_format = '#,##0'
        # EERR acum: lógica de arranque en mes con ingresos — se deja valor calculado en ERP.
        ws.cell(row=r, column=10, value=float(row.get("EERR_ACUM") or 0)).number_format = '#,##0'

    total_row = data_start + n_rows
    tot = fila_total if fila_total is not None else fila_total_flujo_mensual(df_flujo)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)

    first_data = data_start
    last_data = data_start + n_rows - 1
    for col_idx in range(_COL_ING, _COL_EG_PROY + 1):
        letter = get_column_letter(col_idx)
        ws.cell(
            row=total_row,
            column=col_idx,
            value=f"=SUM({letter}{first_data}:{letter}{last_data})",
        ).number_format = '#,##0'

    ws.cell(
        row=total_row,
        column=_COL_EG_TOTAL,
        value=f"=SUM(H{first_data}:H{last_data})",
    ).number_format = '#,##0'
    ws.cell(
        row=total_row,
        column=_COL_RES,
        value=f"=SUM(I{first_data}:I{last_data})",
    ).number_format = '#,##0'
    # EERR acumulado = último mes (no suma lineal).
    ws.cell(
        row=total_row,
        column=10,
        value=float(tot.get("EERR_ACUM", df_flujo["EERR_ACUM"].iloc[-1]) or 0),
    ).number_format = '#,##0'
    ws.cell(row=total_row, column=10).font = Font(bold=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(header_row, total_row + 1):
        for c in range(1, len(_DISPLAY_COLS) + 1):
            ws.cell(row=r, column=c).border = border
            if r >= data_start and c > 1:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(_DISPLAY_COLS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14 if col_idx > 1 else 12

    ws.freeze_panes = ws.cell(row=data_start, column=1)

    nota_row = total_row + 2
    ws.cell(
        row=nota_row,
        column=1,
        value="Nota: EGRESOS TOTAL = EGRESOS REAL + EGRESOS PROY. RESULTADO MES = INGRESOS − EGRESOS TOTAL. "
        "EERR ACUM viene del ERP (arranque en primer mes con ingresos).",
    )
    ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=10)
    ws.cell(row=nota_row, column=1).font = Font(size=9, italic=True, color="666666")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename_flujo_excel(temporada: str) -> str:
    return f"flujo_mensual_{_safe_name(temporada)}.xlsx"
